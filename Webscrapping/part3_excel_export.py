"""
Legal Tech Competitor Research System - Part 3: Excel Export
Takes research results from JSON and creates formatted benchmark Excel file
"""

import json
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from datetime import datetime

class BenchmarkExcelExporter:
    def __init__(self):
        self.research_data = []
        self.benchmark_df = None
        self.output_file = "Benchmark.xlsx"
        
        # Define column mapping and order
        self.column_mapping = {
            'competitor': 'Competitor',
            'website': 'Website',
            'business_model': 'Business Model',
            'ai_powered_legal_chatbot': 'AI-Powered Legal Chatbot',
            'stage': 'Stage',
            'fundraising_stage': 'Fundraising stage',
            'multilingual_support': 'Multilingual Support',
            'mobile_web_accessibility': 'Mobile & Web Accessibility',
            'api_integration': 'API Integration',
            'free_tier': 'Free Tier',
            'subscription_based': 'Subscription-Based',
            'pricing': 'Pricing',
            'target_audience': 'Target Audience',
            'user_base_growth_rate': 'User Base & Growth Rate',
            'partnerships_integrations': 'Partnerships & Integrations',
            'coverage': 'Coverage',
            'product': 'Product',
            'founding_team_rating': 'Founding Team (/5)',
            'direct_indirect': 'Direct / Indirect',
            'comment': 'Comment'
        }

    def create_benchmark_excel(self, research_file="company_research_results.json"):
        """Main method to create formatted Excel benchmark file"""
        print("PART 3: Creating Excel Benchmark File")
        print("=" * 50)
        
        # Load research data
        if not self.load_research_data(research_file):
            return False
        
        # Convert to DataFrame
        self.convert_to_dataframe()
        
        # Create Excel file with formatting
        self.create_formatted_excel()
        
        # Add analysis sheets
        self.add_analysis_sheets()
        
        print(f"SUCCESS: Benchmark Excel file created: {self.output_file}")
        return True

    def load_research_data(self, filename):
        """Load research results from JSON file"""
        try:
            if os.path.exists(filename):
                with open(filename, 'r', encoding='utf-8') as f:
                    self.research_data = json.load(f)
                print(f"SUCCESS: Loaded research data for {len(self.research_data)} companies")
                return True
            else:
                print(f"ERROR: Research file {filename} not found")
                print("Please run Part 2 first to generate research data")
                return False
        except Exception as e:
            print(f"Error loading research data: {e}")
            return False

    def convert_to_dataframe(self):
        """Convert JSON research data to pandas DataFrame"""
        print("Converting research data to DataFrame...")
        
        # Prepare data for DataFrame
        df_data = []
        
        for company in self.research_data:
            row_data = {}
            
            # Map all fields according to column mapping
            for json_key, excel_column in self.column_mapping.items():
                value = company.get(json_key, 'Unknown')
                
                # Clean and format values
                if isinstance(value, dict):
                    # Convert dict to readable string
                    value = self.format_dict_value(value)
                elif isinstance(value, list):
                    # Convert list to comma-separated string
                    value = ', '.join(str(item) for item in value)
                elif value is None or value == '':
                    value = 'Unknown'
                
                row_data[excel_column] = str(value)
            
            df_data.append(row_data)
        
        # Create DataFrame with proper column order
        column_order = list(self.column_mapping.values())
        self.benchmark_df = pd.DataFrame(df_data, columns=column_order)
        
        print(f"SUCCESS: Created DataFrame with {len(self.benchmark_df)} companies and {len(column_order)} columns")

    def format_dict_value(self, value):
        """Format dictionary values for Excel display"""
        if not isinstance(value, dict):
            return str(value)
        
        # Format as key: value pairs
        formatted_items = []
        for k, v in value.items():
            if isinstance(v, (list, dict)):
                v = str(v)
            formatted_items.append(f"{k}: {v}")
        
        return '; '.join(formatted_items)

    def create_formatted_excel(self):
        """Create Excel file with professional formatting"""
        print("Creating formatted Excel file...")
        
        # Create Excel writer
        with pd.ExcelWriter(self.output_file, engine='openpyxl') as writer:
            # Write main benchmark data
            self.benchmark_df.to_excel(writer, sheet_name='Benchmark', index=False)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Benchmark']
            
            # Apply formatting
            self.format_benchmark_sheet(worksheet)

    def format_benchmark_sheet(self, worksheet):
        """Apply professional formatting to the benchmark sheet"""
        print("Applying professional formatting...")
        
        # Define styles
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        data_font = Font(name='Arial', size=10)
        data_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format header row
        for col_num, column in enumerate(self.benchmark_df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border_style
        
        # Format data rows
        for row_num in range(2, len(self.benchmark_df) + 2):
            for col_num in range(1, len(self.benchmark_df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = border_style
                
                # Apply conditional formatting based on content
                self.apply_conditional_formatting(cell, col_num)
        
        # Adjust column widths
        self.adjust_column_widths(worksheet)
        
        # Freeze header row
        worksheet.freeze_panes = 'A2'

    def apply_conditional_formatting(self, cell, col_num):
        """Apply conditional formatting based on cell content"""
        value = str(cell.value).lower()
        
        # Color coding for Direct/Indirect column
        if col_num == len(self.benchmark_df.columns):  # Last column (Direct/Indirect)
            if 'direct' in value:
                cell.fill = PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type='solid')
            elif 'indirect' in value:
                cell.fill = PatternFill(start_color='FFF2E8', end_color='FFF2E8', fill_type='solid')
        
        # Color coding for AI Chatbot column
        ai_chatbot_col = list(self.column_mapping.values()).index('AI-Powered Legal Chatbot') + 1
        if col_num == ai_chatbot_col:
            if 'yes' in value:
                cell.fill = PatternFill(start_color='E8F8E8', end_color='E8F8E8', fill_type='solid')
            elif 'no' in value:
                cell.fill = PatternFill(start_color='F8E8E8', end_color='F8E8E8', fill_type='solid')
        
        # Color coding for funding stage
        stage_col = list(self.column_mapping.values()).index('Stage') + 1
        if col_num == stage_col:
            if any(term in value for term in ['series c', 'established', 'growth']):
                cell.fill = PatternFill(start_color='E8E8F8', end_color='E8E8F8', fill_type='solid')
            elif any(term in value for term in ['series a', 'series b']):
                cell.fill = PatternFill(start_color='F0F0F8', end_color='F0F0F8', fill_type='solid')

    def adjust_column_widths(self, worksheet):
        """Adjust column widths for better readability"""
        column_widths = {
            'Competitor': 20,
            'Website': 25,
            'Business Model': 15,
            'AI-Powered Legal Chatbot': 25,
            'Stage': 15,
            'Fundraising stage': 18,
            'Multilingual Support': 20,
            'Mobile & Web Accessibility': 25,
            'API Integration': 15,
            'Free Tier': 12,
            'Subscription-Based': 18,
            'Pricing': 20,
            'Target Audience': 20,
            'User Base & Growth Rate': 25,
            'Partnerships & Integrations': 30,
            'Coverage': 15,
            'Product': 35,
            'Founding Team (/5)': 18,
            'Direct / Indirect': 18,
            'Comment': 30
        }
        
        for col_num, column in enumerate(self.benchmark_df.columns, 1):
            width = column_widths.get(column, 15)
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width

    def add_analysis_sheets(self):
        """Add analysis sheets to the Excel file"""
        print("Adding analysis sheets...")
        
        # Load the workbook to add sheets
        workbook = openpyxl.load_workbook(self.output_file)
        
        # Add summary analysis sheet
        self.create_summary_sheet(workbook)
        
        # Add competitive positioning sheet
        self.create_positioning_sheet(workbook)
        
        # Add German market focus sheet
        self.create_german_market_sheet(workbook)
        
        # Save the workbook
        workbook.save(self.output_file)

    def create_summary_sheet(self, workbook):
        """Create summary analysis sheet"""
        summary_sheet = workbook.create_sheet("Summary Analysis")
        
        # Summary statistics
        total_companies = len(self.benchmark_df)
        direct_competitors = len(self.benchmark_df[self.benchmark_df['Direct / Indirect'].str.contains('Direct', na=False)])
        ai_chatbot_companies = len(self.benchmark_df[self.benchmark_df['AI-Powered Legal Chatbot'].str.contains('Yes', na=False)])
        german_coverage = len(self.benchmark_df[self.benchmark_df['Coverage'].str.contains('German', na=False)])
        
        summary_data = [
            ['Metric', 'Value', 'Percentage'],
            ['Total Companies Analyzed', total_companies, '100%'],
            ['Direct Competitors', direct_competitors, f'{direct_competitors/total_companies*100:.1f}%'],
            ['Companies with AI Chatbots', ai_chatbot_companies, f'{ai_chatbot_companies/total_companies*100:.1f}%'],
            ['Companies with German Coverage', german_coverage, f'{german_coverage/total_companies*100:.1f}%'],
            [''],
            ['Analysis Date', datetime.now().strftime('%Y-%m-%d %H:%M'), ''],
            ['Data Source', 'Multi-agent research system', ''],
            ['Research Method', 'Web scraping + LLM analysis', '']
        ]
        
        for row_num, row_data in enumerate(summary_data, 1):
            for col_num, value in enumerate(row_data, 1):
                cell = summary_sheet.cell(row=row_num, column=col_num, value=value)
                if row_num == 1:  # Header
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                    cell.font = Font(bold=True, color='FFFFFF')

    def create_positioning_sheet(self, workbook):
        """Create competitive positioning analysis sheet"""
        positioning_sheet = workbook.create_sheet("Competitive Positioning")
        
        # Analyze competitive positioning
        direct_competitors = self.benchmark_df[self.benchmark_df['Direct / Indirect'].str.contains('Direct', na=False)]
        
        positioning_data = [
            ['Company', 'AI Chatbot', 'German Market', 'Funding Stage', 'Business Model', 'Competitive Threat'],
            ['', '', '', '', '', '']
        ]
        
        for _, company in direct_competitors.iterrows():
            threat_level = self.assess_competitive_threat(company)
            positioning_data.append([
                company['Competitor'],
                company['AI-Powered Legal Chatbot'],
                'Yes' if 'German' in str(company['Coverage']) else 'No',
                company['Stage'],
                company['Business Model'],
                threat_level
            ])
        
        for row_num, row_data in enumerate(positioning_data, 1):
            for col_num, value in enumerate(row_data, 1):
                cell = positioning_sheet.cell(row=row_num, column=col_num, value=value)
                if row_num == 1:  # Header
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                    cell.font = Font(bold=True, color='FFFFFF')

    def assess_competitive_threat(self, company):
        """Assess competitive threat level for a company"""
        threat_score = 0
        
        # AI Chatbot capability
        if 'yes' in str(company['AI-Powered Legal Chatbot']).lower():
            threat_score += 3
        
        # German market presence
        if 'german' in str(company['Coverage']).lower():
            threat_score += 3
        
        # Funding stage
        stage = str(company['Stage']).lower()
        if any(term in stage for term in ['series c', 'established']):
            threat_score += 2
        elif any(term in stage for term in ['series a', 'series b']):
            threat_score += 1
        
        # Business model similarity
        if 'saas' in str(company['Business Model']).lower():
            threat_score += 1
        
        # Determine threat level
        if threat_score >= 7:
            return 'High'
        elif threat_score >= 4:
            return 'Medium'
        else:
            return 'Low'

    def create_german_market_sheet(self, workbook):
        """Create German market focus analysis sheet"""
        german_sheet = workbook.create_sheet("German Market Focus")
        
        # Filter companies with German market presence
        german_companies = self.benchmark_df[
            self.benchmark_df['Coverage'].str.contains('German|Germany', na=False, case=False) |
            self.benchmark_df['Multilingual Support'].str.contains('German', na=False, case=False)
        ]
        
        german_data = [
            ['Company', 'Coverage', 'Multilingual Support', 'AI Chatbot', 'Pricing', 'Oratio Relevance'],
            ['', '', '', '', '', '']
        ]
        
        for _, company in german_companies.iterrows():
            relevance = self.assess_oratio_relevance(company)
            german_data.append([
                company['Competitor'],
                company['Coverage'],
                company['Multilingual Support'],
                company['AI-Powered Legal Chatbot'],
                company['Pricing'],
                relevance
            ])
        
        for row_num, row_data in enumerate(german_data, 1):
            for col_num, value in enumerate(row_data, 1):
                cell = german_sheet.cell(row=row_num, column=col_num, value=value)
                if row_num == 1:  # Header
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                    cell.font = Font(bold=True, color='FFFFFF')

    def assess_oratio_relevance(self, company):
        """Assess relevance to Oratio Technologies"""
        relevance_score = 0
        
        # German market presence
        if 'german' in str(company['Coverage']).lower():
            relevance_score += 3
        
        # AI legal chatbot
        if 'yes' in str(company['AI-Powered Legal Chatbot']).lower():
            relevance_score += 3
        
        # Individual/consumer focus
        if 'individual' in str(company['Target Audience']).lower():
            relevance_score += 2
        
        # Multilingual German support
        if 'german' in str(company['Multilingual Support']).lower():
            relevance_score += 1
        
        if relevance_score >= 7:
            return 'Very High'
        elif relevance_score >= 5:
            return 'High'
        elif relevance_score >= 3:
            return 'Medium'
        else:
            return 'Low'

def run_part_3(research_file="company_research_results.json", output_file="Benchmark.xlsx"):
    """Run Part 3: Excel Export"""
    exporter = BenchmarkExcelExporter()
    exporter.output_file = output_file
    success = exporter.create_benchmark_excel(research_file)
    return success

if __name__ == "__main__":
    run_part_3()